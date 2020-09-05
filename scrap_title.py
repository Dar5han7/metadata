from bs4 import BeautifulSoup
from requests import get
from openpyxl import load_workbook, Workbook
from time import sleep
from random import randint
import csv
import os

outfile = open("AB_MUE_03.09.2020_01.csv", "w", newline='',encoding="utf-8")
writer = csv.writer(outfile)

# wb_url_list = load_workbook('nvli_url_list.xlsm')
# ws_url_list = wb_url_list.active
# url_list = [ws_url_list.cell(row=i, column=6).value for i in range(491, 505)]
# website_list = []

# excel_file_name = 'AB_MUE_20-08-2020.xlsx'
file_extn_list =[]
img_path = r"F:\nvli\AB_MUE_03.09.2020_ANAND"
for file in os.listdir(img_path):
    # print(file)
    file_extn_list.append(file)

website_list = []
for file_extn in file_extn_list:
    t = 'http://museumsofindia.gov.in/repository/record/'+file_extn
    website_list.append(t)

# for i in range(0, len(url_list)):
#     # print(i)
#     accession_number = url_list[i].split('/')
#     t = 'http://museumsofindia.gov.in/repository/record/'+accession_number[-2]
#     website_list.append(t)

wb = Workbook()
ws = wb.active

# ws.append(website_list)
count=1
new_list = website_list[134:]

for url in new_list:
    print(url)
    htmlString = get(url).text
    sleep(2)
    html = BeautifulSoup(htmlString, 'lxml')
    table_tag = html.select("table")[0]
    tab_data = [[item.text for item in row_data.select("th,td")]
                for row_data in table_tag.select("tr")]

    for data in tab_data:
        writer.writerow(data)
        print(' '.join(data))