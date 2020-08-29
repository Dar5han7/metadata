import autoit
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException
import urllib, os, urllib.request
import time

from openpyxl import *
from collections import defaultdict
from selenium.webdriver.support.ui import Select


driver = webdriver.Chrome()
driver.implicitly_wait(10)
#
usrName = 'anand.swaroop.nvli@gmail.com'
pssWrd = "anand@1#3"

driver.maximize_window()
driver.get("http://moirepository.nvli.in/password-login")

driver.find_element_by_name('login_email').send_keys(usrName)
driver.find_element_by_name('login_password').send_keys(pssWrd)
driver.find_element_by_name('login_submit').click()
wait = WebDriverWait(driver, 2)
time.sleep(5)

# print(driver.page_source)
driver.find_element_by_name('submit_own').click()
# time.sleep(10)

all_handler_details = {}
excel_file_name = 'edited_entries.xlsx'
try:
    wb = load_workbook(excel_file_name)
except:
    wb = Workbook()

ws = wb.active
ws["A1"] = "Handler No"


for i in range(900,1218):

    driver.find_element_by_xpath('//*[@id="content"]/div[2]/table/tbody/tr['+str(i)+']/td[2]/a').click()
    time.sleep(1)
    driver.find_element_by_xpath('/html/body/main/div[3]/div/div[2]/div/div[2]/form[1]/input[2]').click()
    entry = driver.find_element_by_name('value_dc_subject_01').text
    check = driver.find_element_by_name('value_dc_subject_00').text
    if entry != "Artwork" and check =="Miniature Painting":
        time.sleep(1)
        driver.find_element_by_name('value_dc_subject_01').clear()
        driver.find_element_by_name('value_dc_subject_01').send_keys("Artwork")
        time.sleep(1)
        accesion_number = driver.find_element_by_name('value_dc_identifier_accessionnumber_00').text
        ws[f"A{i}"] = accesion_number
    else:
      pass
    driver.find_element_by_xpath('/html/body/main/div[2]/form/div[3]/input[5]').click()
    time.sleep(1)
    driver.find_element_by_xpath('/html/body/header/div/nav/div/ul/li/a').click()
    time.sleep(1)
    driver.find_element_by_xpath('/html/body/header/div/nav/div/ul/li/ul/li[2]/a').click()
    time.sleep(1)
    driver.find_element_by_name('submit_own').click()
    time.sleep(1)

#     accesion_number = driver.find_element_by_xpath('//*[@id="content"]/div[3]/div/div[1]/div[2]/table/tbody/tr[2]/td[1]/a').text


#
wb.save(excel_file_name)
wb.close()