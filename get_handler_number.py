import autoit
from selenium import webdriver
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException
import urllib, os, urllib.request
import time

from openpyxl import *
from collections import defaultdict
from selenium.webdriver.support.ui import Select


driver = webdriver.Chrome(ChromeDriverManager().install())
driver.implicitly_wait(10)
#
usrName = 'anand.swaroop.nvli@gmail.com'
pssWrd = "Anand@Swaroop$NVLI@!"

driver.maximize_window()
driver.get("http://moirepository.nvli.in/password-login")

driver.find_element_by_name('login_email').send_keys(usrName)
driver.find_element_by_name('login_password').send_keys(pssWrd)
driver.find_element_by_name('login_submit').click()
wait = WebDriverWait(driver, 2)
time.sleep(5)

# print(driver.page_source)
driver.find_element_by_name('submit_own').click()
time.sleep(10)

all_handler_details = {}
excel_file_name = 'handler-1-1000.xlsx'
try:
    wb = load_workbook(excel_file_name)
except:
    wb = Workbook()



ws = wb.active
ws["A1"] = "Handler No"
ws["B1"] = "Accession No"
ws["C1"] = "index No"
ws["D1"] = "href"




for i in range(6550,6623):
    # // *[ @ id = "content"] / div[3] / div / div[1] / table / tbody / tr[6] / td[2] / a
    time.sleep(1)
    driver.find_element_by_xpath('//*[@id="content"]/div[2]/table/tbody/tr['+str(i)+']/td[2]/a').click()
    time.sleep(1)
    # try://*[@id="content"]/div[3]/div/div[1]/div[2]/table/tbody/tr[6]/td[2]
    #     handle_number = driver.find_element_by_xpath('// *[ @ id = "content"] / div[3] / div / div[1] / table / tbody / tr[6] / td[2] / a').text
    #
    # except Exception:
    #     handle_number = driver.find_element_by_xpath('//*[@id="content"]/div[3]/div/div[1]/table/tbody/tr[5]/td[2]/a').text
    time.sleep(2)
    handle_number = driver.find_element_by_xpath('// *[ @ id = "content"] / div[3] / div / div[1] / div[1] / code').text
    time.sleep(2)

    accesion_number = driver.find_element_by_xpath('//*[@id="content"]/div[3]/div/div[1]/div[2]/table/tbody/tr[2]/td[1]/a').text
    ws[f"A{i}"] = handle_number
    ws[f"B{i}"] = accesion_number
    ws[f"C{i}"] = i
    ws[f"D{i}"] = '//*[@id="content"]/div[2]/table/tbody/tr['+str(i)+']/td[2]/a'
    print(i,handle_number,accesion_number)
    driver.back()

    wb.save(excel_file_name)
wb.close()

