from bs4 import BeautifulSoup
from requests import get
import requests
from bs4 import BeautifulSoup
import html
from openpyxl import load_workbook, Workbook
import time
import random
import getpass
import os
from pathlib import Path

# ####### Define for your user names ##############
# url = 'http://datastore.nvli.in/yash/alh_ald/yash/'
# user, password = 'yash','Yash@$#1290'
# image_folder_name = 'Yash_images'
# excel_file_name = 'yash_excel.xlsx'
# # start 0 and end 10 gives images and data in first 10 files
# start = 9
# end = 12
#
# user_name = getpass.getuser()
# if user_name == 'Deepthi':
#     url = 'http://datastore.nvli.in/deepthi/alh_ald/DEEPTHI_04.08.2020/'
#     user, password = 'deepthi', 'Deepthi@Reddy#341'
#     image_folder_name = 'BEADS_401_900_images'
#     excel_file_name = 'raw_extract_07_08.xlsx'
#     # start 0 and end 10 gives images and data in first 10 files
#     print('its me')
#     start = 0
#     end = 2
#
# ##############
# # This routine reads urls from datastore and gets file extentions from start and end specified
# ##############
# resp = requests.get(url, auth=(user, password))
# print(resp.content)
# fragment = BeautifulSoup(resp.content,features='lxml')
# url_list = []
# all_index_files = list(fragment.findAll('a'))
# if start is None:
#     start = 0
# if end is None:
#     end = len(all_index_files)
# newlist = all_index_files[start+5:end+5]
#
# file_extn_list = []
# for file_name in newlist:
#     file_extn = file_name.string
#     file_extn_list.append(file_extn)
# ###############################################################
# # Scraping images from datastore
# ##############
# # this routine downloads images for all file extentions created
# ##############
#
# for file_extn in file_extn_list:
#     file_url = url+file_extn
#     resp = requests.get(file_url,auth=(user, password))
#     file_name = file_extn.split('/')[0]
#     fragment = BeautifulSoup(resp.content,features='lxml')
#     file_data = fragment.findAll('a')
#     file_data_string = []
#     for x in file_data:
#         file_data_string.append(x.string)
#
#     for link in file_data_string:
#         if '_l.jpg' in link:
#             print(link.rsplit('_', 1)[0])
#             if link.rsplit('_', 1)[0] + '_h.jpg' in file_data_string:
#                 continue
#             else:
#                 img_url = file_url + link
#                 resp = requests.get(img_url, auth=(user, password))
#
#                 Path(image_folder_name + '/' + file_name).mkdir(parents=True, exist_ok=True)
#                 with open(image_folder_name + '/' + file_name + '/'+link, 'wb') as f:
#                     f.write(resp.content)
#         elif '_h.jpg' in link:
#             img_url = file_url + link
#             resp = requests.get(img_url, auth=(user, password))
#             Path(image_folder_name+'/'+file_name).mkdir(parents=True, exist_ok=True)
#             with open(image_folder_name+'/'+file_name+'/'+link, 'wb') as f:
#                 f.write(resp.content)

################################################################
# Scraping text from MOI website
# ##########################
# This routine gets data from museums of india website
# ##########################
excel_file_name = 'extractAB_MUE_20-08-2020.xlsx'
file_extn_list =[]
img_path = r"F:\nvli\AB_MUE_20-08-2020"
for file in os.listdir(img_path):
    # print(file)
    file_extn_list.append(file)


website_list = []
for file_extn in file_extn_list:
    t = 'http://museumsofindia.gov.in/repository/record/'+file_extn
    website_list.append(t)

# ###########
# Comment this for irregular data
# ###########
try:
    wb = load_workbook(excel_file_name)
except:
    wb = Workbook()
ws = wb.active
for web_url in website_list:
    for file in os.listdir(img_path):
        for data in os.listdir((img_path+"\\"+file)):
            if data.endswith('_dc.xml'):
                xml_url = data

    xml_url = url + web_url.split('/')[-2] + '/' + web_url.split('/')[-2] + '_dc.xml'
    time.sleep(2)
    htmlString = get(web_url).text
    html = BeautifulSoup(htmlString, 'lxml')
    entries = html.find_all('td')
    text = [e.get_text() for e in entries]
    print('{} posts were found.'.format(len(text)))
    text.append(web_url)
    text.append(xml_url)
    ws.append(text)
    wb.save(excel_file_name)
wb.close()


# ###############
# # Uncomment this for irregular data
# ################
# import csv
# outfile = open("table_data1.csv", "w", newline='')
# writer = csv.writer(outfile)
# count = 1
# for url in website_list:
#     htmlString = get(url).text
#     time.sleep(random.randint(15, 20))
#     html = BeautifulSoup(htmlString, 'lxml')
#     table_tag = html.select("table")[0]
#     tab_data = [[item.text for item in row_data.select("th,td")]
#                 for row_data in table_tag.select("tr")]
#
#     for data in tab_data:
#         writer.writerow(data)
#         print(' '.join(data))
# outfile.close()
#
# input_csv_file_name = "table_data1.csv"
# output_csv_file_name = "output5.csv"
# title_word = "Title"
# tables = []
#
# with open(input_csv_file_name, 'r') as csv_file:
#     csv_reader = csv.reader(csv_file, delimiter=',')
#     table = []
#     for row in csv_reader:
#         if row[0] == title_word:
#             tables.append(table)
#             table = []
#         table.append(row)
#     tables.append(table)
#
# tables = tables[1:]
# for table in tables:
#     for row in table:
#         print(row)
#     print()
#
# number_of_tables = len(tables)
# parsed = {}
#
# for table_number, table in enumerate(tables):
#     for row in table:
#         key = row[0]
#         value = row[1]
#         try:
#             parsed[key][table_number] = value
#         except:
#             parsed[key] = ["" for x in range(number_of_tables)]
#             parsed[key][table_number] = value
#
# outfile = open(output_csv_file_name, "w", newline='')
# csv_writer = csv.writer(outfile)
# for key, value in parsed.items():
#     row = [key] + value
#     print(row)
#     csv_writer.writerow(row)